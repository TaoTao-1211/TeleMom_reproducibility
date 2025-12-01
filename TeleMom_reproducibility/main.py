import os
import json
from openai import OpenAI
from openpyxl import load_workbook
import re
from volcenginesdkarkruntime import Ark
from bert_score import BERTScorer
wb_question= load_workbook(r"C:\Users\zhutao\Desktop\TeleMom_question.xlsx")
wb_answer = load_workbook(r"C:\Users\zhutao\Desktop\TeleMom_answer.xlsx")
wb_adjudicate = load_workbook(r"C:\Users\zhutao\Desktop\TeleMom_adjudicate.xlsx")

ws_question = wb_question['Sheet1']
ws_answer = wb_answer['Sheet1']
ws_adjudicate = wb_adjudicate['Sheet1']

pattern = r'\{[^}]*\}'#正则表达式，以确保输出格式为JSON

# #-------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
model_sep=1#确定当前模型选择顺序，以便后续写入

#先调用deepssek-chatAPI（非思考模式）
for row_idx, row in enumerate(ws_question.iter_rows(min_row=2, values_only=True), start=2):
    type=row[0]#查看问题类型
    if type=='query':#问题是询问的形式
        try:
            client = OpenAI(
                api_key="sk-71b898da1d6746ff9f127c79d92261ac",
                base_url="https://api.deepseek.com")
            messages = [
                {"role": "system",
                 "content": "You are an expert in an telecommunication technical committee. Your role is to give suggestion to the adjudicator who make final decisions."},
                {"role": "user",
                 "content": "Please provide the answers to the following telecommunications related questions. The questions will be in a JSON format, the answers must also be in a JSON format as follows strictly:{\"Answer\": \"***\"}"},
                {"role": "user", "content": "{\"Question\":\" " + row[1] + "\"}"}
            ]
            response = client.chat.completions.create(
                model="deepseek-chat",
                messages=messages,
                stream=False,
                )
            for round in range(4):
                response1 = client.chat.completions.create(
                    model="deepseek-chat",
                    messages=[
                        {"role": "system","content": "You are an expert in an telecommunication technical committee. Your role is to give suggestion to the adjudicator who make final decisions."},
                        {"role": "user","content": "Please provide a confidence score for the answer based on the question and the given answer. The score should be a natural integer between 0 and 10. If the score is greater than or equal to 7, the answer is considered highly credible. The questions will be in a JSON format, the answers must also be in a JSON format as follows strictly:{\"score\": \"Specific score\"}"},
                        {"role": "user", "content": "{\"Question\":\" " + row[1] + "\",\""+re.match(pattern, response.choices[0].message.content).group()+"\"}"}
                    ],
                    stream=False
                )
                score = json.loads(re.match(pattern, response1.choices[0].message.content).group())
                if int(score["score"]) >=7:
                    answer = json.loads(re.match(pattern, response.choices[0].message.content).group())
                    ws_answer.cell(row=row_idx, column=model_sep*2-1, value=answer["Answer"])
                    wb_answer.save(r"C:\Users\zhutao\Desktop\TeleMom_answer.xlsx")
                    break
                else:
                    messages.append(response.choices[0].message)
                    messages.append({"role": "user", "content": "Due to low confidence, please re-enter the output.the answers must also be in a JSON format as follows strictly:{\"Answer\": \"***\"}"})
                    response = client.chat.completions.create(
                        model="deepseek-chat",
                        messages=messages,
                        stream=False,
                        temperature = 0.3,
                        presence_penalty = 0.1,
                        frequency_penalty = 0.1
                    )


        except Exception as e:
            # 捕获异常并记录错误，但不中断循环
            error_msg = f"API调用失败: {str(e)}"
            print(f"deepssek-chatAPI第 {row_idx} 行处理失败: {error_msg}")
            ws_answer.cell(row=row_idx, column=model_sep, value=error_msg)
            # 即使出错也保存，确保错误信息被记录
            wb_answer.save(r"C:\Users\zhutao\Desktop\TeleMom_answer.xlsx")





    if type=='option':#问题是可供选择的形式
        try:
            client = OpenAI(
                api_key="sk-71b898da1d6746ff9f127c79d92261ac",
                base_url="https://api.deepseek.com")
            messages = [
                {"role": "system",
                 "content": "You are an expert in an telecommunication technical committee. Your role is to give suggestion to the adjudicator who make final decisions."},
                {"role": "user",
                 "content": "Please provide the answers to the following telecommunications related questions. The questions will be in a JSON format, the answers must also be in a JSON format as follows strictly:{\"Answer\": \"这里填写具体的选项如Option1\"}，{\"Reason\": \"这里填写原因解释\"}"},
                {"role": "user",
                 "content": "{\"Question\":\" " + row[1] + "\"},{\"Option1\":\" " + row[2] + "\"},{\"Option2\":\" " +
                            row[3] + "\"},{\"Option3\":\" " + row[4] + "\"},{\"Option4\":\" " + row[5] + "\"}"}
            ]
            response = client.chat.completions.create(
                model="deepseek-chat",
                messages=messages,
                stream=False
                )
            for round in range(4):
                response1 = client.chat.completions.create(
                    model="deepseek-chat",
                    messages=[
                        {"role": "system",
                         "content": "You are an expert in an telecommunication technical committee. Your role is to give suggestion to the adjudicator who make final decisions."},
                        {"role": "user",
                         "content": "Please provide a confidence score for the answer based on the question and the given answer. The score should be a natural integer between 0 and 10. If the score is greater than or equal to 7, the answer is considered highly credible. The questions will be in a JSON format, the answers must also be in a JSON format as follows strictly:{\"score\": \"Specific score\"}"},
                        {"role": "user", "content": "{\"Question\":\" " + row[1] + "\",\"" + re.match(pattern,response.choices[0].message.content).group() + "\"}"}
                    ],
                    stream=False
                )
                score = json.loads(re.match(pattern, response1.choices[0].message.content).group())
                if int(score["score"]) >= 7:
                    answer = json.loads(re.match(pattern, response.choices[0].message.content).group())
                    ws_answer.cell(row=row_idx, column=model_sep * 2 - 1, value=answer["Answer"])
                    wb_answer.save(r"C:\Users\zhutao\Desktop\TeleMom_answer.xlsx")
                    break
                else:
                    messages.append(response.choices[0].message)
                    messages.append({"role": "user",
                                     "content": "Due to low confidence, please re-enter the output.the answers must also be in a JSON format as follows strictly:{\"Answer\": \"这里填写具体的选项如Option1\"}，{\"Reason\": \"这里填写原因解释\"}"})
                    response = client.chat.completions.create(
                        model="deepseek-chat",
                        messages=messages,
                        stream=False,
                        temperature=0.3,
                        presence_penalty=0.1,
                        frequency_penalty=0.1
                    )
        except Exception as e:
            # 捕获异常并记录错误，但不中断循环
            error_msg = f"API调用失败: {str(e)}"
            print(f"deepssek-chatAPI第 {row_idx} 行处理失败: {error_msg}")
            ws_answer.cell(row=row_idx, column=1, value=error_msg)
            # 即使出错也保存，确保错误信息被记录
            wb_answer.save(r"C:\Users\zhutao\Desktop\TeleMom_answer.xlsx")



# #-------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------



model_sep=2#确定当前模型选择顺序，以便后续写入

#先调用deepseek-reasoner（思考模式）
for row_idx, row in enumerate(ws_question.iter_rows(min_row=2, values_only=True), start=2):
    type=row[0]#查看问题类型
    if type=='query':#问题是询问的形式
        try:
            client = OpenAI(
                api_key="sk-71b898da1d6746ff9f127c79d92261ac",
                base_url="https://api.deepseek.com")
            messages = [
                {"role": "system",
                 "content": "You are an expert in an telecommunication technical committee. Your role is to give suggestion to the adjudicator who make final decisions."},
                {"role": "user",
                 "content": "Please provide the answers to the following telecommunications related questions. The questions will be in a JSON format, the answers must also be in a JSON format as follows strictly:{\"Answer\": \"***\"}"},
                {"role": "user", "content": "{\"Question\":\" " + row[1] + "\"}"}
            ]
            response = client.chat.completions.create(
                model="deepseek-reasoner",
                messages=messages,
                stream=False
                )
            for round in range(4):
                response1 = client.chat.completions.create(
                    model="deepseek-reasoner",
                    messages=[
                        {"role": "system",
                         "content": "You are an expert in an telecommunication technical committee. Your role is to give suggestion to the adjudicator who make final decisions."},
                        {"role": "user",
                         "content": "Please provide a confidence score for the answer based on the question and the given answer. The score should be a natural integer between 0 and 10. If the score is greater than or equal to 7, the answer is considered highly credible. The questions will be in a JSON format, the answers must also be in a JSON format as follows strictly:{\"score\": \"Specific score\"}"},
                        {"role": "user", "content": "{\"Question\":\" " + row[1] + "\",\"" + re.match(pattern,response.choices[0].message.content).group() + "\"}"}
                    ],
                    stream=False
                )
                score = json.loads(re.match(pattern, response1.choices[0].message.content).group())
                if int(score["score"]) >= 7:
                    answer = json.loads(re.match(pattern, response.choices[0].message.content).group())
                    ws_answer.cell(row=row_idx, column=model_sep * 2 - 1, value=answer["Answer"])
                    wb_answer.save(r"C:\Users\zhutao\Desktop\TeleMom_answer.xlsx")
                    break
                else:
                    messages.append(response.choices[0].message)
                    messages.append({"role": "user",
                                     "content": "Due to low confidence, please re-enter the output.the answers must also be in a JSON format as follows strictly:{\"Answer\": \"***\"}"})
                    response = client.chat.completions.create(
                        model="deepseek-reasoner",
                        messages=messages,
                        stream=False,
                        temperature=0.3,
                        presence_penalty=0.1,
                        frequency_penalty=0.1
                    )


        except Exception as e:
            # 捕获异常并记录错误，但不中断循环
            error_msg = f"API调用失败: {str(e)}"
            print(f"deepseek-reasoner第 {row_idx} 行处理失败: {error_msg}")
            ws_answer.cell(row=row_idx, column=model_sep, value=error_msg)
            # 即使出错也保存，确保错误信息被记录
            wb_answer.save(r"C:\Users\zhutao\Desktop\TeleMom_answer.xlsx")

    if type=='option':#问题是可供选择的形式
        try:
            client = OpenAI(
                api_key="sk-71b898da1d6746ff9f127c79d92261ac",
                base_url="https://api.deepseek.com")

            messages = [
                {"role": "system",
                 "content": "You are an expert in an telecommunication technical committee. Your role is to give suggestion to the adjudicator who make final decisions."},
                {"role": "user",
                 "content": "Please provide the answers to the following telecommunications related questions. The questions will be in a JSON format, the answers must also be in a JSON format as follows strictly:{\"Answer\": \"这里填写具体的选项如Option1\"}，{\"Reason\": \"这里填写原因解释\"}"},
                {"role": "user",
                 "content": "{\"Question\":\" " + row[1] + "\"},{\"Option1\":\" " + row[2] + "\"},{\"Option2\":\" " +
                            row[3] + "\"},{\"Option3\":\" " + row[4] + "\"},{\"Option4\":\" " + row[5] + "\"}"}
            ]
            response = client.chat.completions.create(
                model="deepseek-reasoner",
                messages=messages,
                stream=False
            )
            for round in range(4):
                response1 = client.chat.completions.create(
                    model="deepseek-reasoner",
                    messages=[
                        {"role": "system",
                         "content": "You are an expert in an telecommunication technical committee. Your role is to give suggestion to the adjudicator who make final decisions."},
                        {"role": "user",
                         "content": "Please provide a confidence score for the answer based on the question and the given answer. The score should be a natural integer between 0 and 10. If the score is greater than or equal to 7, the answer is considered highly credible. The questions will be in a JSON format, the answers must also be in a JSON format as follows strictly:{\"score\": \"Specific score\"}"},
                        {"role": "user", "content": "{\"Question\":\" " + row[1] + "\",\"" + re.match(pattern,response.choices[0].message.content).group() + "\"}"}
                    ],
                    stream=False
                )
                score = json.loads(re.match(pattern, response1.choices[0].message.content).group())
                if int(score["score"]) >= 7:
                    answer = json.loads(re.match(pattern, response.choices[0].message.content).group())
                    ws_answer.cell(row=row_idx, column=model_sep * 2 - 1, value=answer["Answer"])
                    wb_answer.save(r"C:\Users\zhutao\Desktop\TeleMom_answer.xlsx")
                    break
                else:
                    messages.append(response.choices[0].message)
                    messages.append({"role": "user",
                                     "content": "Due to low confidence, please re-enter the output.the answers must also be in a JSON format as follows strictly:{\"Answer\": \"Fill in specific options here, such as Option1\"}，{\"Reason\": \"Fill in the reason explanation here\"}"})
                    response = client.chat.completions.create(
                        model="deepseek-reasoner",
                        messages=messages,
                        stream=False,
                        temperature=0.3,
                        presence_penalty=0.1,
                        frequency_penalty=0.1
                    )
        except Exception as e:
            # 捕获异常并记录错误，但不中断循环
            error_msg = f"API调用失败: {str(e)}"
            print(f"deepseek-reasoner第 {row_idx} 行处理失败: {error_msg}")
            ws_answer.cell(row=row_idx, column=1, value=error_msg)
            # 即使出错也保存，确保错误信息被记录
            wb_answer.save(r"C:\Users\zhutao\Desktop\TeleMom_answer.xlsx")

# # ----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

model_sep = 3  # 确定当前模型选择顺序，以便后续写入

# 调用doubao-seed-1-6-251015
for row_idx, row in enumerate(ws_question.iter_rows(min_row=2, values_only=True), start=2):
    type = row[0]  # 查看问题类型
    if type == 'query':  # 问题是询问的形式
        try:
            client = OpenAI(
                # 此为默认路径，您可根据业务所在地域进行配置
                base_url="https://ark.cn-beijing.volces.com/api/v3",
                # 从环境变量中获取您的 API Key。此为默认方式，您可根据需要进行修改
                api_key="24dc010a-2688-4c5d-a062-c6fff61291e7",
            )
            messages = [
                {"role": "system",
                 "content": "You are an expert in an telecommunication technical committee. Your role is to give suggestion to the adjudicator who make final decisions."},
                {"role": "user",
                 "content": "Please provide the answers to the following telecommunications related questions. The questions will be in a JSON format, the answers must also be in a JSON format as follows strictly:{\"Answer\": \"***\"}"},
                {"role": "user", "content": "{\"Question\":\" " + row[1] + "\"}"}
            ]
            response = client.chat.completions.create(
                model="doubao-seed-1-6-251015",
                messages=messages,
                stream=False
            )
            for round in range(4):
                response1 = client.chat.completions.create(
                    model="doubao-seed-1-6-251015",
                    messages=[
                        {"role": "system",
                         "content": "You are an expert in an telecommunication technical committee. Your role is to give suggestion to the adjudicator who make final decisions."},
                        {"role": "user",
                         "content": "Please provide a confidence score for the answer based on the question and the given answer. The score should be a natural integer between 0 and 10. If the score is greater than or equal to 7, the answer is considered highly credible. The questions will be in a JSON format, the answers must also be in a JSON format as follows strictly:{\"score\": \"Specific score\"}"},
                        {"role": "user", "content": "{\"Question\":\" " + row[1] + "\",\"" + re.match(pattern,response.choices[0].message.content).group() + "\"}"}
                    ],
                    stream=False
                )
                score = json.loads(re.match(pattern, response1.choices[0].message.content).group())
                if int(score["score"]) >= 7:
                    answer = json.loads(re.match(pattern, response.choices[0].message.content).group())
                    ws_answer.cell(row=row_idx, column=model_sep * 2 - 1, value=answer["Answer"])
                    wb_answer.save(r"C:\Users\zhutao\Desktop\TeleMom_answer.xlsx")
                    break
                else:
                    messages.append(response.choices[0].message)
                    messages.append({"role": "user",
                                     "content": "Due to low confidence, please re-enter the output.the answers must also be in a JSON format as follows strictly:{\"Answer\": \"***\"}"})
                    response = client.chat.completions.create(
                        model="doubao-seed-1-6-251015",
                        messages=messages,
                        stream=False,
                        temperature=0.3,
                        presence_penalty=0.1,
                        frequency_penalty=0.1
                    )


        except Exception as e:
            # 捕获异常并记录错误，但不中断循环
            error_msg = f"API调用失败: {str(e)}"
            print(f"doubao-seed-1-6-251015第 {row_idx} 行处理失败: {error_msg}")
            ws_answer.cell(row=row_idx, column=model_sep, value=error_msg)
            # 即使出错也保存，确保错误信息被记录
            wb_answer.save(r"C:\Users\zhutao\Desktop\TeleMom_answer.xlsx")

    if type == 'option':  # 问题是可供选择的形式
        try:
            client = Ark(
                api_key="24dc010a-2688-4c5d-a062-c6fff61291e7",
                # The base URL for model invocation .
                base_url="https://ark.cn-beijing.volces.com/api/v3", )
            messages = [
                {"role": "system",
                 "content": "You are an expert in an telecommunication technical committee. Your role is to give suggestion to the adjudicator who make final decisions."},
                {"role": "user",
                 "content": "Please provide the answers to the following telecommunications related questions. The questions will be in a JSON format, the answers must also be in a JSON format as follows strictly:{\"Answer\": \"这里填写具体的选项如Option1\"}，{\"Reason\": \"这里填写原因解释\"}"},
                {"role": "user",
                 "content": "{\"Question\":\" " + row[1] + "\"},{\"Option1\":\" " + row[2] + "\"},{\"Option2\":\" " +
                            row[3] + "\"},{\"Option3\":\" " + row[4] + "\"},{\"Option4\":\" " + row[5] + "\"}"}
            ]
            response = client.chat.completions.create(
                model="doubao-seed-1-6-251015",
                messages=messages,
                stream=False
            )
            for round in range(4):
                response1 = client.chat.completions.create(
                    model="doubao-seed-1-6-251015",
                    messages=[
                        {"role": "system",
                         "content": "You are an expert in an telecommunication technical committee. Your role is to give suggestion to the adjudicator who make final decisions."},
                        {"role": "user",
                         "content": "Please provide a confidence score for the answer based on the question and the given answer. The score should be a natural integer between 0 and 10. If the score is greater than or equal to 7, the answer is considered highly credible. The questions will be in a JSON format, the answers must also be in a JSON format as follows strictly:{\"score\": \"Specific score\"}"},
                        {"role": "user", "content": "{\"Question\":\" " + row[1] + "\",\"" + re.match(pattern,response.choices[0].message.content).group() + "\"}"}
                    ],
                    stream=False
                )
                score = json.loads(re.match(pattern, response1.choices[0].message.content).group())
                if int(score["score"]) >= 7:
                    answer = json.loads(re.match(pattern, response.choices[0].message.content).group())
                    ws_answer.cell(row=row_idx, column=model_sep * 2 - 1, value=answer["Answer"])
                    wb_answer.save(r"C:\Users\zhutao\Desktop\TeleMom_answer.xlsx")
                    break
                else:
                    messages.append(response.choices[0].message)
                    messages.append({"role": "user",
                                     "content": "Due to low confidence, please re-enter the output.the answers must also be in a JSON format as follows strictly:{\"Answer\": \"Fill in specific options here, such as Option1\"}，{\"Reason\": \"Fill in the reason explanation here\"}"})
                    response = client.chat.completions.create(
                        model="doubao-seed-1-6-251015",
                        messages=messages,
                        stream=False,
                        temperature=0.3,
                        presence_penalty=0.1,
                        frequency_penalty=0.1
                    )
        except Exception as e:
            # 捕获异常并记录错误，但不中断循环
            error_msg = f"API调用失败: {str(e)}"
            print(f"doubao-seed-1-6-251015第 {row_idx} 行处理失败: {error_msg}")
            ws_answer.cell(row=row_idx, column=1, value=error_msg)
            # 即使出错也保存，确保错误信息被记录
            wb_answer.save(r"C:\Users\zhutao\Desktop\TeleMom_answer.xlsx")



#-------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------


# 以下是Adjudicator的内容

for row_idx, row in enumerate(ws_answer.iter_rows(min_row=2, values_only=True), start=2):
    try:
        client = OpenAI(
            api_key="sk-68120db749ba4b8ebc7f9bd2be095e49",
            base_url="https://dashscope.aliyuncs.com/compatible-mode/v1",
        )
        messages = [
            {"role": "system",
             "content": "You are an expert in an telecommunication technical committee. Your role is to give suggestion to the adjudicator who make final decisions."},
            {"role": "user",
             "content": "Based on the information given below, answers the question in the telecommunication field.{\"Question\":\" " + ws_question.cell(row=row_idx, column=1).value + "\","
                        "\"model 1 name\":\"deepseek-chat\",\"Answer\":\"" + row[0] + "\",\"Reason\":\"" + str(row[1]) + "\",\"model 2 name\":\"deepseek-reasoner\","
                        "\"Answer\":\"" + row[2] + "\",\"Reason\":\"" + str(row[3]) + "\",\"model 3 name\":\"doubao-1.6\",\"Answer\":\"" + row[4] + "\",\"Reason\":\"" + str(row[5]) + "\"}"},
            {"role": "user", "content": "Analyse the information given, and give your answer. Respond in JSON with the following structure strictly:{\"Answer\":\"Final answer\",\"Reason\":"
                                        "\"If the preceding Reson is not NULL, summarize and output Reson; if it is None, output None.\"}"}
        ]
        response = client.chat.completions.create(
            model="qwen3-max",
            messages=messages,
            stream=False
        )
        answer = json.loads(re.match(pattern, response.choices[0].message.content).group())
        response1 = client.chat.completions.create(
                model="qwen3-max",
                messages=[
                            {"role": "system",
                             "content": "You are an expert in an telecommunication technical committee. Your role is to give suggestion to the adjudicator who make final decisions."},
                            {"role": "user",
                             "content": "Please provide a confidence score for the answer based on the question and the given answer. The score should be a natural integer between 0 and 10. If the score is greater than or equal to 7, the answer is considered highly credible. The questions will be in a JSON format, the answers must also be in a JSON format as follows strictly:{\"score\": \"Specific score\"}"},
                            {"role": "user", "content": "{\"Question\":\" " + ws_question.cell(row=row_idx, column=1).value + "\",\"" + re.match(pattern,response.choices[0].message.content).group() + "\"}"}
                        ],
                stream=False
                        )
        score = json.loads(re.match(pattern, response1.choices[0].message.content).group())
        if int(score["score"]) >= 7:
            ws_adjudicate.cell(row=row_idx, column=1, value=answer["Answer"])
            ws_adjudicate.cell(row=row_idx, column=2, value=answer["Reason"])
            ws_adjudicate.cell(row=row_idx, column=3, value="High confidence")
            wb_adjudicate.save(r"C:\Users\zhutao\Desktop\TeleMom_adjudicate.xlsx")
        else:
            ws_adjudicate.cell(row=row_idx, column=1, value=answer["Answer"])
            ws_adjudicate.cell(row=row_idx, column=2, value=answer["Reason"])
            ws_adjudicate.cell(row=row_idx, column=3, value="Low confidence, requires manual verification.")
            wb_adjudicate.save(r"C:\Users\zhutao\Desktop\TeleMom_adjudicate.xlsx")

    except Exception as e:
        # 捕获异常并记录错误，但不中断循环
        error_msg = f"API调用失败: {str(e)}"
        print(f"第 {row_idx} 行处理失败: {error_msg}")
        ws_adjudicate.cell(row=row_idx, column=1, value=error_msg)
        # 即使出错也保存，确保错误信息被记录
        wb_adjudicate.save(r"C:\Users\zhutao\Desktop\TeleMom_adjudicate.xlsx")


# ------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------


#评分系统设置

model_path = r"C:\Users\zhutao\PycharmProjects\TeleMom_reproducibility\models\bert-base-chinese"
candidates = []
column_letter = 3

# 遍历从第2行到最大行
for row in ws_question.iter_rows(min_row=2, values_only=True):
    if len(row) > 2 and row[2] is not None and str(row[2]).strip():
        candidates.append(str(row[2]).strip())
references = []
column_letter = 1

# 遍历从第2行到最大行
for row in ws_adjudicate.iter_rows(min_row=2, values_only=True):
    # 检查该行是否有数据（例如第一个单元格不为空）
    if row[0] is not None and str(row[0]).strip():
        references.append(str(row[0]).strip())
# 使用 BERTScorer 类，可以接受本地路径
print(candidates)
scorer = BERTScorer(
    model_type=model_path,
    lang="zh",
    num_layers=12  # bert-base-chinese 有12层
)

P, R, F1 = scorer.score(candidates, references)
for i in range(len(F1)):
    print(f"BERTScore F1: {F1[i].item():.4f}")


